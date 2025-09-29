# -*- coding: utf-8 -*-
import subprocess
import re
import openpyxl
import os

def check_adb_device():
    """
    检查是否有ADB设备连接。
    """
    try:
        result = subprocess.run(['adb', 'devices'], capture_output=True, text=True, check=True, encoding='utf-8')
        # 'List of devices attached' 后面如果有多于一行，说明有设备
        lines = result.stdout.strip().splitlines()
        if len(lines) > 1:
            print("ADB device found.")
            return True
        else:
            print("Error: No ADB device connected. Please connect a device and enable USB debugging.")
            return False
    except (subprocess.CalledProcessError, FileNotFoundError):
        print("Error: 'adb' command not found. Please ensure ADB is installed and in your system's PATH.")
        return False

def get_packages():
    """
    获取所有第三方应用的包名列表。
    使用 'pm list packages -3' 来过滤掉系统应用，如果你需要所有应用，可以改成 'pm list packages'。
    """
    print("Fetching installed packages...")
    try:
        # 为了进行全面排查，我们获取所有应用包名
        result = subprocess.run(['adb', 'shell', 'pm', 'list', 'packages'], capture_output=True, text=True, check=True, encoding='utf-8')
        packages = [line.replace('package:', '').strip() for line in result.stdout.strip().splitlines()]
        print(f"Found {len(packages)} packages.")
        return packages
    except subprocess.CalledProcessError as e:
        print(f"Error getting package list: {e}")
        return []

def get_package_info(package_name):
    """
    获取单个应用的权限、用户ID和用户组ID。
    (已根据Android 14+的格式反馈进行更新)
    """
    print(f"Analyzing package: {package_name}...")
    try:
        # 执行 dumpsys package 命令获取详细信息
        result = subprocess.run(['adb', 'shell', 'dumpsys', 'package', package_name], capture_output=True, text=True, check=True, encoding='utf-8')
        output = result.stdout
        
        # 兼容新版 Android (appId) 和旧版 (userId)
        user_id = "Not Found"
        # 优先匹配 appId
        app_id_match = re.search(r'^\s*appId=(\d+)', output, re.MULTILINE)
        if app_id_match:
            user_id = app_id_match.group(1)
        else:
            # 回退匹配 userId
            user_id_match = re.search(r'^\s*userId=(\d+)', output, re.MULTILINE)
            if user_id_match:
                user_id = user_id_match.group(1)

        # 提取 gids (在某些新版系统中可能不显示)
        gids_match = re.search(r'^\s*gids=\[(.*?)\]', output, re.MULTILINE)
        gids_str = gids_match.group(1) if gids_match else ""
        gids = [gid.strip() for gid in gids_str.split(',') if gid.strip()]

        # 提取申请的权限 (requested permissions)，使用更稳定的正则
        permissions = []
        # 匹配 "requested permissions:" 标题行，并捕获后续所有缩进的行
        perm_match = re.search(r'^\s*requested permissions:\s*\n((?:\s+.*\n?)*)', output, re.MULTILINE)
        if perm_match:
            permissions_block = perm_match.group(1).strip()
            # 按行分割并去除空行
            permissions = [p.strip() for p in permissions_block.splitlines() if p.strip()]

        return {
            'user_id': user_id,
            'gids': gids,
            'permissions': permissions
        }

    except subprocess.CalledProcessError as e:
        print(f"Could not get info for {package_name}: {e}")
        return None
    except Exception as e:
        print(f"An unexpected error occurred while processing {package_name}: {e}")
        return None


def create_excel_report(all_package_data, filename="android_permissions_audit.xlsx"):
    """
    将所有应用的数据写入一个Excel文件，每个应用一个sheet。
    """
    print(f"Creating Excel report: {filename}...")
    workbook = openpyxl.Workbook()
    # 删除默认创建的sheet
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    for package_name, data in all_package_data.items():
        if not data:
            continue

        # Excel sheet名称有31个字符的限制，并且不能包含某些特殊字符
        safe_sheet_name = re.sub(r'[\\/*?:\[\]]', '_', package_name)
        if len(safe_sheet_name) > 31:
            safe_sheet_name = safe_sheet_name[-31:] # 取包名的后31位，更有辨识度

        sheet = workbook.create_sheet(title=safe_sheet_name)

        # 写入表头
        header = ['Type', 'Detail', 'Reason for Application (Please fill in)']
        sheet.append(header)
        
        # 设置列宽
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 60
        sheet.column_dimensions['C'].width = 50


        # 写入用户信息
        sheet.append(['User ID / App ID', data['user_id'], ''])
        sheet.append(['Group IDs', ', '.join(data['gids']) if data['gids'] else 'Not Found', ''])
        
        # 写入一个空行作为分隔
        sheet.append([]) 
        
        # 写入权限信息
        if data['permissions']:
            sheet.append(['Permission', '', '']) # 权限小标题
            for perm in sorted(data['permissions']):
                sheet.append(['', perm, ''])
        else:
            sheet.append(['Permission', 'No permissions requested.', ''])


    try:
        workbook.save(filename)
        print(f"Successfully saved report to: {os.path.abspath(filename)}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")

def main():
    """
    主函数
    """
    if not check_adb_device():
        return

    packages = get_packages()
    if not packages:
        print("No packages found to analyze. Exiting.")
        return

    all_data = {}
    for pkg in packages:
        info = get_package_info(pkg)
        if info and (info['user_id'] != "Not Found" or info['gids'] or info['permissions']):
             all_data[pkg] = info
        else:
             print(f"Warning: No valid data found for package {pkg}. It might be a system component with no dumpable info.")

    
    if all_data:
        create_excel_report(all_data)
    else:
        print("Could not retrieve information for any package.")


if __name__ == "__main__":
    main()

